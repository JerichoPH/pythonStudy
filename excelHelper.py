import argparse
import os
import sys
from typing import Any, List, Dict, Tuple

import openpyxl


class ExcelReader:
    _excel: openpyxl.workbook.Workbook = None
    _sheet: openpyxl.worksheet.worksheet.Worksheet = None
    _content: list[Any] = []
    _rows: list = []
    _title: list[str] = []
    _filename: str = ''
    _read_title_row_number: int = 1
    _original_row_number: int = 2
    _finished_row_number: int = None

    def __init__(self, filename: str, read_title_row_number: int = None, original_row_number: int = None, finished_row_number: int = None):
        """
        初始化
        :param filename: 文件名
        :type filename: str
        :param read_title_row_number: 设置读取表头行标
        :type read_title_row_number: int
        :param original_row_number: 设置读取起始行标
        :type original_row_number: int
        :param finished_row_number: 设置读取终止行标
        :type finished_row_number: int
        """
        self._filename = filename
        if read_title_row_number is not None:
            self._read_title_row_number = read_title_row_number
        if original_row_number is not None:
            self._original_row_number = original_row_number
        if finished_row_number is not None:
            self._finished_row_number = finished_row_number

    def __enter__(self) -> __init__:
        self._excel = openpyxl.load_workbook(self._filename)
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self._excel.close()

    def open_excel(self, filename: str) -> __init__:
        """
        打开Excel文件
        :param filename: 文件路径
        :type filename: str
        :return: 本类对象
        :rtype: excelReader.ExcelReader.ExcelReader
        """
        self._excel = openpyxl.load_workbook(filename)
        return self

    def read_title(self) -> __init__:
        """
        读取一行数据
        :return: 本类对象
        :rtype: excelReader.ExcelReader.ExcelReader
        """
        self._title = [str(cell.value) for cell in tuple(self._sheet.rows)[self.get_read_title_row_number]]
        return self

    def read_rows(self) -> __init__:
        """
        读取多行数据
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        self._rows = list(self._sheet.rows)[self.get_original_row_number:self.get_finished_row_number if self.get_finished_row_number else None]
        self._content = [[str(cell.value) for cell in row_datum] for row_datum in tuple(self._sheet.rows)[self.get_original_row_number - 1:self.get_finished_row_number if self.get_finished_row_number else None]]
        return self

    def read_entire_sheet_by_first(self) -> __init__:
        """
        读取第一个sheet
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        self.set_sheet_by_index().read_title().read_rows()
        return self

    def read_entire_sheet_by_active(self) -> __init__:
        """
        读取激活的sheet
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        self.set_sheet(self._excel.active).read_title().read_rows()
        return self

    def read_entire_sheet_by_name(self, worksheet_name: str) -> __init__:
        """
        根据sheet名称读取
        :param worksheet_name: sheet名称
        :type worksheet_name: str
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        self.set_sheet_by_name(worksheet_name).read_title().read_rows()
        return self

    @property
    def get_excel(self) -> openpyxl.workbook.Workbook:
        """
        获取excel对象
        :return: Excel对象
        :rtype: openpyxl.workbook.Workbook
        """
        return self._excel

    @property
    def get_sheet(self) -> openpyxl.worksheet.worksheet.Worksheet:
        """
        获取sheet
        :return: Worksheet对象
        :rtype: openpyxl.worksheet.worksheet.Worksheet
        """
        return self._sheet

    def set_sheet(self, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> __init__:
        """
        设置sheet
        :param worksheet: sheet对象
        :type worksheet: openpyxl.worksheet.worksheet.Worksheet
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        self._sheet = worksheet
        return self

    def set_sheet_by_index(self, worksheet_index: int = 0) -> __init__:
        """
        通过索引设置工作表
        :param worksheet_index: 工作表索引
        :type worksheet_index: int
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        self._sheet = self._excel[self._excel.sheetnames[worksheet_index]]
        return self

    def set_sheet_by_active(self) -> __init__:
        """
        根据当前激活状态设置工作表
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        self._sheet = self._excel.active
        return self

    def set_sheet_by_name(self, worksheet_name: str) -> __init__:
        """
        通过名称设置工作表
        :param worksheet_name: 工作表名称
        :type worksheet_name: str
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        self._sheet = self._excel[worksheet_name]
        return self

    @property
    def get_title(self) -> list:
        """
        获取表头
        :return: 表头
        :rtype: tuple
        """
        return self._title

    def set_title(self, titles: list) -> __init__:
        """
        设置表头
        :param titles: 表头
        :type titles: tuple
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        self._title = titles
        return self

    @property
    def get_rows(self) -> list:
        """
        获取原始行
        :return: 原始行列表
        :rtype: list
        """
        return self._rows

    @property
    def get_read_title_row_number(self) -> int:
        """
        获取读取表头行标
        :return: 行标
        :rtype: int
        """
        return self._read_title_row_number if self._read_title_row_number else 0

    def set_read_title_row_number(self, row_number: int) -> __init__:
        """
        设置读取表头行
        :param row_number: 行标
        :type row_number: int
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        self._read_title_row_number = row_number
        return self

    @property
    def get_original_row_number(self) -> int:
        """
        获取起始读取行标
        :return: 行标
        :rtype: int
        """
        return self._original_row_number if self._original_row_number else 1

    def set_original_row_number(self, now_number: int) -> __init__:
        """
        设置起始读取行标
        :param now_number: 行标
        :type now_number: int
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        self._original_row_number = now_number
        return self

    @property
    def get_finished_row_number(self) -> int:
        return self._finished_row_number if self._finished_row_number else 0

    def set_finished_row_number(self, row_number: int) -> __init__:
        """
        设置终止读取行标
        :param row_number: 行标
        :type row_number: int
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        self._finished_row_number = row_number + 1
        return self

    def save(self, filename: str = None) -> __init__:
        """
        保存文件
        :param filename: 文件名（如不填写则保存回原文件）
        :type filename: str
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        self._excel.save(filename if filename else self._filename)
        return self

    @property
    def get_row_dimensions(self) -> Dict[str, openpyxl.worksheet.dimensions.RowDimension]:
        """
        获取所有行高
        :return: 全表行高
        :rtype: dict
        """
        return self._sheet.row_dimensions

    def set_row_dimensions(self, row_index: int, row_height: float) -> __init__:
        """
        设置行高
        :param row_index: 行索引
        :type row_index: int
        :param row_height: 行高（0～409，默认12.75）
        :type row_height: float
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        self._sheet[row_index].height = row_height
        return self

    @property
    def get_column_dimensions(self) -> List[Tuple[str, openpyxl.worksheet.dimensions.ColumnDimension]]:
        """
        获取列宽
        :return: 全表列宽
        :rtype: list
        """
        return self._sheet.column_dimensions

    def set_column_dimensions(self, column_index: str, column_width: float) -> __init__:
        """
        设置列宽
        :param column_index: 列坐标
        :type column_index: str
        :param column_width: 列宽（0～255，默认8.43个字符）
        :type column_width: float
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        self._sheet[column_index].width = column_width
        return self

    def set_merge_cells(self, original_cell_coordinate: str, finished_cell_coordinate: str) -> __init__:
        """
        设置合并单元格
        :param original_cell_coordinate: 起始坐标
        :type original_cell_coordinate: str
        :param finished_cell_coordinate: 终止坐标
        :type finished_cell_coordinate: str
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        self._sheet.merge_cells(f'{original_cell_coordinate}:{finished_cell_coordinate}')
        return self

    def set_unmerge_cells(self, original_cell_coordinate: str, finished_cell_coordinate: str) -> __init__:
        """
        拆分单元格
        :param original_cell_coordinate: 起始坐标
        :type original_cell_coordinate: str
        :param finished_cell_coordinate: 终止坐标
        :type finished_cell_coordinate: str
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        self._sheet.unmerge_cells(f'{original_cell_coordinate}:{finished_cell_coordinate}')
        return self

    def set_freeze_panes(self, cell_coordinate: str) -> __init__:
        """
        设置冻结窗口
        :param cell_coordinate: 单元格坐标（A1：冻结首行，B1：冻结首列，B2：冻结首行和首列）
        :type cell_coordinate: str
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        self._sheet.freeze_panes = cell_coordinate
        return self

    def add_chart_bar(self, original_row: int, original_col: int, finished_row: int, finished_col: int, chart_target_coordinate: str, chart_title: str = None, x_axis_title: str = None, y_axis_title: str = None):
        """
        添加条形图
        :param original_row:
        :type original_row:
        :param original_col:
        :type original_col:
        :param finished_row:
        :type finished_row:
        :param finished_col:
        :type finished_col:
        :param chart_target_coordinate:
        :type chart_target_coordinate:
        :param chart_title:
        :type chart_title:
        :param x_axis_title:
        :type x_axis_title:
        :param y_axis_title:
        :type y_axis_title:
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        values = openpyxl.chart.Reference(self._sheet, min_row=original_row, min_col=original_col, max_row=finished_row, max_col=finished_col)
        chart = openpyxl.chart.BarChart()
        if chart_title is not None:
            chart.title = chart_title
        if x_axis_title is not None:
            chart.x_axis.title = x_axis_title
        if y_axis_title is not None:
            chart.y_axis.title = y_axis_title
        chart.add_data(values)
        self._sheet.add_chart(chart, chart_target_coordinate)
        return self

    def add_chart_line(self, original_row: int, original_col: int, finished_row: int, finished_col: int, chart_target_coordinate: str, chart_title: str = None, x_axis_title: str = None, y_axis_title: str = None):
        """
        设置折线图
        :param original_row:
        :type original_row:
        :param original_col:
        :type original_col:
        :param finished_row:
        :type finished_row:
        :param finished_col:
        :type finished_col:
        :param chart_target_coordinate:
        :type chart_target_coordinate:
        :param chart_title:
        :type chart_title:
        :param x_axis_title:
        :type x_axis_title:
        :param y_axis_title:
        :type y_axis_title:
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        values = openpyxl.chart.Reference(self._sheet, min_row=original_row, min_col=original_col, max_row=finished_row, max_col=finished_col)
        chart = openpyxl.chart.LineChart()
        if chart_title is not None:
            chart.title = chart_title
        if x_axis_title is not None:
            chart.x_axis.title = x_axis_title
        if y_axis_title is not None:
            chart.y_axis.title = y_axis_title
        chart.add_data(values)
        self._sheet.add_chart(chart, chart_target_coordinate)
        return self

    def add_chart_scatter(self, original_row: int, original_col: int, finished_row: int, finished_col: int, chart_target_coordinate: str, chart_title: str = None, x_axis_title: str = None, y_axis_title: str = None):
        """
        添加散点图
        :param original_row:
        :type original_row:
        :param original_col:
        :type original_col:
        :param finished_row:
        :type finished_row:
        :param finished_col:
        :type finished_col:
        :param chart_target_coordinate:
        :type chart_target_coordinate:
        :param chart_title:
        :type chart_title:
        :param x_axis_title:
        :type x_axis_title:
        :param y_axis_title:
        :type y_axis_title:
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        values = openpyxl.chart.Reference(self._sheet, min_row=original_row, min_col=original_col, max_row=finished_row, max_col=finished_col)
        chart = openpyxl.chart.ScatterChart()
        if chart_title is not None:
            chart.title = chart_title
        if x_axis_title is not None:
            chart.x_axis.title = x_axis_title
        if y_axis_title is not None:
            chart.y_axis.title = y_axis_title
        chart.add_data(values)
        self._sheet.add_chart(chart, chart_target_coordinate)
        return self

    def add_chart_pie(self, original_row: int, original_col: int, finished_row: int, finished_col: int, chart_target_coordinate: str, chart_title: str = None, x_axis_title: str = None, y_axis_title: str = None):
        """
        添加饼图
        :param original_row:
        :type original_row:
        :param original_col:
        :type original_col:
        :param finished_row:
        :type finished_row:
        :param finished_col:
        :type finished_col:
        :param chart_target_coordinate:
        :type chart_target_coordinate:
        :param chart_title:
        :type chart_title:
        :param x_axis_title:
        :type x_axis_title:
        :param y_axis_title:
        :type y_axis_title:
        :return: 本类对象
        :rtype: excelHelper.ExcelReader
        """
        values = openpyxl.chart.Reference(self._sheet, min_row=original_row, min_col=original_col, max_row=finished_row, max_col=finished_col)
        chart = openpyxl.chart.PieChart()
        if chart_title is not None:
            chart.title = chart_title
        if x_axis_title is not None:
            chart.x_axis.title = x_axis_title
        if y_axis_title is not None:
            chart.y_axis.title = y_axis_title
        chart.add_data(values)
        self._sheet.add_chart(chart, chart_target_coordinate)
        return self

    @property
    def to_dict(self) -> dict:
        """
        返回字典格式数据（必须设置表头）
        :return: Excel数据（带表头）
        :rtype: dict
        """
        if self._title:
            contents = [dict(zip(self._title, content)) for content in self._content]
            return {str(row_number + self.get_original_row_number): row_datum for row_number, row_datum in enumerate(contents)}
        else:
            return {}

    @property
    def to_list(self) -> list:
        """
        返回数组格式数据
        :return:
        :rtype:
        """
        return self._content


class ExcelWriterCell:
    _content: str = ''
    _coordinate: str = None
    _font_name: str = None
    _font_color: str = None
    _font_size: float = None
    _font_underline: str = None
    _font_bold: bool = None
    _font_italic: bool = None
    _font_vert_align: str = None
    _font_outline: bool = None
    _font_shadow: bool = None
    _border_style: str = None
    _border_color: str = '000000'
    _border_top_style: str = None
    _border_top_color: str = '000000'
    _border_bottom_style: str = None
    _border_bottom_color: str = '000000'
    _border_left_style: str = None
    _border_left_color: str = '000000'
    _border_right_style: str = None
    _border_right_color: str = '000000'
    _fill_fg_color: str = None
    _alignment_horizontal = 'left'
    _alignment_vertical = 'center'

    ALIGNMENT_HORIZONTAL_LEFT = 'left'
    ALIGNMENT_HORIZONTAL_RIGHT = 'right'
    ALIGNMENT_HORIZONTAL_CENTER = 'center'
    ALIGNMENT_VERTICAL_TOP = 'top'
    ALIGNMENT_VERTICAL_BOTTOM = 'bottom'
    ALIGNMENT_VERTICAL_CENTER = 'center'

    BORDER_STYLE_THIN = 'thin'
    BORDER_STYLE_DOTTED = 'dotted'
    BORDER_STYLE_THICK = 'thick'
    BORDER_STYLE_MEDIUM_DASHED = 'mediumDashed'
    BORDER_STYLE_DASH_DOT_DOT = 'dashDotDot'
    BORDER_STYLE_MEDIUM = 'medium'
    BORDER_STYLE_DASH_DOT = 'dashDot'
    BORDER_STYLE_SLANT_DASH_DOT = 'slantDashDot'
    BORDER_STYLE_DASHED = 'dashed',
    BORDER_STYLE_MEDIUM_DASH_DOT = 'mediumDashDot'
    BORDER_STYLE_MEDIUM_DASH_DOT_DOT = 'mediumDashDotDot'
    BORDER_STYLE_HAIR = 'hair'
    BORDER_STYLE_DOUBLE = 'double'

    def __init__(
            self,
            content: str = '',
            coordinate: str = None,
            font_name: str = None,
            font_color: str = None,
            font_size: float = None,
            font_underline: str = None,
            font_bold: bool = None,
            font_italic: bool = None,
            font_vert_align: str = None,
            font_outline: bool = None,
            font_shadow: bool = None,
            border_style: str = None,
            border_color: str = None,
            border_top_style: str = None,
            border_top_color: str = None,
            border_bottom_style: str = None,
            border_bottom_color: str = None,
            border_left_style: str = None,
            border_left_color: str = None,
            border_right_style: str = None,
            border_right_color: str = None,
            fill_fg_color: str = None,
            alignment_horizontal: str = None,
            alignment_vertical: str = None,
    ):
        """
        初始化
        :param content: 单元格内容
        :type content: str
        :param coordinate: 单元格所在坐标
        :type coordinate: str
        :param font_name: 字体
        :type font_name: str
        :param font_color: 颜色
        :type font_color: str
        :param font_size: 字号
        :type font_size: float
        :param font_underline: 下划线
        :type font_underline: str
        :param font_bold: 粗体
        :type font_bold: bool
        :param font_italic: 斜体
        :type font_italic: bool
        :param font_vert_align: 对齐
        :type font_vert_align: str
        :param font_outline: 外框线
        :type font_outline: bool
        :param font_shadow: 阴影
        :type font_shadow: bool
        :param border_style: 边框样式
        :type border_style: str
        :param border_top_style: 上边框样式
        :type border_top_style: str
        :param border_bottom_style: 下边框样式
        :type border_bottom_style: str
        :param border_left_style: 左边框样式
        :type border_left_style: str
        :param border_right_style: 右边框样式
        :type border_right_style: str
        :param fill_fg_color: 填充背景色
        :type fill_fg_color: str
        :param alignment_horizontal: 字体水平剧中
        :type alignment_horizontal: str
        :param alignment_vertical: 字体垂直居中
        :type alignment_vertical: str
        """
        self._content = content
        self._coordinate = coordinate
        if font_name is not None:
            self._font_name = font_name
        if font_color is not None:
            self._font_color = font_color
        if font_size is not None:
            self._font_size = font_size
        if font_underline is not None:
            self._font_underline = font_underline
        if font_bold is not None:
            self._font_bold = font_bold
        if font_italic is not None:
            self._font_italic = font_italic
        if font_vert_align is not None:
            self._font_vert_align = font_vert_align
        if font_outline is not None:
            self._font_outline = font_outline
        if font_shadow is not None:
            self._font_shadow = font_shadow
        if border_style is not None:
            self._border_style = border_style
        if border_color is not None:
            self._border_color = border_color
        if border_top_style is not None:
            self._border_top_style = border_top_style
        if border_top_color is not None:
            self._border_top_color = border_top_color
        if border_bottom_style is not None:
            self._border_bottom_style = border_bottom_style
        if border_bottom_color is not None:
            self._border_bottom_color = border_bottom_color
        if border_left_style is not None:
            self._border_left_style = border_left_style
        if border_left_color is not None:
            self._border_left_color = border_left_color
        if border_right_style is not None:
            self._border_right_style = border_right_style
        if border_right_color is not None:
            self._border_right_color = border_right_color
        if fill_fg_color is not None:
            self._fill_fg_color = fill_fg_color
        if alignment_horizontal is not None:
            self._alignment_horizontal = alignment_horizontal
        if alignment_vertical is not None:
            self._alignment_vertical = alignment_vertical

    @property
    def get_content(self) -> str:
        """
        获取单元格内容
        :return:
        :rtype:
        """
        return self._content

    def set_content(self, content: str = '') -> __init__:
        """
        设置单元格内容
        :param content: 单元格内容
        :type content:
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """
        self._content = content
        return self

    @property
    def get_coordinate(self) -> str:
        """
        返回单元格所在位置
        :return: 单元格坐标
        :rtype: str
        """
        return self._coordinate

    def set_coordinate(self, coordinate: str) -> __init__:
        """
        设置单元格所在位置
        :param coordinate: 单元格坐标
        :type coordinate: str
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """
        self._coordinate = coordinate
        return self

    @property
    def get_font_name(self) -> str:
        """
        获取当前字体
        :return: 字体名称
        :rtype: str
        """
        return self._font_name

    def set_font_name(self, font_name: str, when: bool = True) -> __init__:
        """
        设置字体
        :param font_name: 字体名称
        :type font_name: str
        :param when: 是否执行设置：True
        :type when: bool
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """
        if when is True:
            self._font_name = font_name
        return self

    @property
    def get_font_color(self) -> str:
        """
        获取字体颜色
        :return: 字体颜色
        :rtype: str
        """
        return self._font_color

    def set_font_color(self, font_color: str, when: bool = True) -> __init__:
        """
        设置字体颜色
        :param font_color: 字体颜色
        :type font_color: str
        :param when: 是否设置：True
        :type when: bool
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """
        if when is True:
            self._font_color = font_color
        return self

    @property
    def get_font_size(self) -> float:
        """
        获取字号
        :return: 字号
        :rtype: float
        """
        return self._font_size

    def set_font_size(self, font_size: float, when: bool = True) -> __init__:
        """
        设置字号
        :param font_size: 字号
        :type font_size: float
        :param when: 是否立即设置：True
        :type when: bool
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """
        if when is True:
            self._font_size = font_size
        return self

    @property
    def get_font_underline(self) -> str:
        """
        获取字体下划线
        :return: 下划线
        :rtype: bool
        """
        return self._font_underline

    def set_font_underline(self, font_underline: str, when: bool = True) -> __init__:
        """
        设置字体下划线
        :param font_underline: 下划线
        :type font_underline: bool
        :param when: 是否立即设置：True
        :type when: bool
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """
        if when is True:
            self._font_underline = font_underline
        return self

    @property
    def get_font_bold(self) -> bool:
        """
        获取字体粗体值
        :return: 字体粗体值
        :rtype: int
        """
        return self._font_bold

    def set_font_bold(self, font_bold: bool, when: bool = True) -> __init__:
        """
        设置字体粗体值
        :param font_bold: 粗体值
        :type font_bold:
        :param when: 是否立即设置：True
        :type when: bool
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """
        if when is True:
            self._font_bold = font_bold
        return self

    @property
    def get_font_italic(self) -> bool:
        """
        获取字体倾斜
        :return: 字体是否倾斜
        :rtype: bool
        """
        return self._font_italic

    def set_font_italic(self, font_italic: bool, when: bool = True) -> __init__:
        """
        设置字体是否倾斜
        :param font_italic: 字体是否倾斜
        :type font_italic: bool
        :param when: 是否立即设置：True
        :type when: bool
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """
        if when is True:
            self._font_italic = font_italic
        return self

    @property
    def get_font_vert_align(self) -> str:
        """
        获取字体对其方式
        :return: 字体对其方式
        :rtype: str
        """
        return self._font_vert_align

    def set_font_vert_align(self, font_vert_align: str, when: bool = True) -> __init__:
        """
        设置字体对其方式
        :param font_vert_align: 字体对齐方式
        :type font_vert_align: str
        :param when: 是否立即设置：True
        :type when: bool
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """
        if when is True:
            self._font_vert_align = font_vert_align
        return self

    @property
    def get_font_outline(self) -> bool:
        """
        获取外边框线
        :return: 外边框线
        :rtype: bool
        """
        return self._font_outline

    def set_font_outline(self, font_outline: bool, when: bool = True) -> __init__:
        """
        设置字体外框线
        :param font_outline: 字体外框线
        :type font_outline: bool
        :param when: 是否立即设置：True
        :type when: bool
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """
        if when is True:
            self._font_outline = font_outline
        return self

    @property
    def get_font_shadow(self) -> bool:
        """
        获取字体阴影
        :return: 字体阴影
        :rtype: bool
        """
        return self._font_shadow

    def set_font_shadow(self, font_shadow: bool, when: bool = True) -> __init__:
        """
        设置字体阴影
        :param font_shadow: 字体阴影
        :type font_shadow: bool
        :param when: 是否立即设置：True
        :type when: bool
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """

    @property
    def get_font(self) -> openpyxl.styles.Font:
        """
        获取字体样式
        :return: 字体样式
        :rtype: openpyxl.styles.Font
        """
        return openpyxl.styles.Font(
            name=self.get_font_name,
            color=self.get_font_color,
            size=self.get_font_size,
            underline=self.get_font_underline,
            bold=self.get_font_bold,
            italic=self.get_font_italic,
            vertAlign=self.get_font_vert_align,
            outline=self.get_font_outline,
            shadow=self.get_font_shadow,
        )

    @property
    def get_border_style(self) -> str:
        """
        获取边框样式
        :return: 边框样式
        :rtype: str
        """
        return self._border_style

    def set_border_style(self, border_style: str) -> __init__:
        """
        设置边框样式
        :param border_style: 边框样式
        :type border_style: str
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """
        self._border_style = border_style
        return self

    @property
    def get_border_top_style(self) -> str:
        """
        获取上边框样式
        :return: 上边框样式
        :rtype: str
        """
        return self._border_top_style

    def set_border_top_style(self, border_top_style: str) -> __init__:
        """
        设置上边框样式
        :param border_top_style: 上边框样式
        :type border_top_style: str
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """
        self._border_top_style = border_top_style
        return self

    @property
    def get_border_bottom_style(self) -> str:
        """
        获取下边框样式
        :return: 下边框样式
        :rtype: str
        """
        return self._border_bottom_style

    def set_border_bottom_style(self, border_bottom_style: str) -> __init__:
        """
        设置下边框样式
        :param border_bottom_style: 下边框样式
        :type border_bottom_style: str
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """
        self._border_bottom_style = border_bottom_style
        return self

    @property
    def get_border_left_style(self) -> str:
        """
        获取左边框样式
        :return: 左边框样式
        :rtype: str
        """
        return self._border_left_style

    def set_border_left_style(self, border_left_style: str) -> __init__:
        """
        设置左边框样式
        :param border_left_style: 左边框样式
        :type border_left_style: str
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """
        self._border_left_style = border_left_style
        return self

    @property
    def get_border_right_style(self) -> str:
        """
        获取右边框样式
        :return: 右边框样式
        :rtype: str
        """
        return self._border_right_style

    def set_border_right_style(self, border_right_style: str) -> __init__:
        """
        设置右边框样式
        :param border_right_style: 右边框样式
        :type border_right_style: str
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """
        self._border_right_style = border_right_style
        return self

    @property
    def get_border_color(self) -> str:
        """
        获取边框颜色
        :return: 边框颜色
        :rtype: str
        """
        return self._border_style

    def set_border_color(self, border_color: str) -> __init__:
        """
        设置边框颜色
        :param border_color: 边框颜色
        :type border_color: str
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """
        self._border_color = border_color
        return self

    @property
    def get_border_top_color(self) -> str:
        """
        获取上边框颜色
        :return: 上边框颜色
        :rtype: str
        """
        return self._border_top_color

    def set_border_top_color(self, border_top_color: str) -> __init__:
        """
        设置上边框颜色
        :param border_top_color: 上边框颜色
        :type border_top_color: str
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """
        self._border_top_color = border_top_color
        return self

    @property
    def get_border_bottom_color(self) -> str:
        """
        获取下边框颜色
        :return: 下边框颜色
        :rtype: str
        """
        return self._border_bottom_color

    def set_border_bottom_color(self, border_bottom_color: str) -> __init__:
        """
        设置下边框颜色
        :param border_bottom_color: 下边框颜色
        :type border_bottom_color: str
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """
        self._border_bottom_color = border_bottom_color
        return self

    @property
    def get_border_left_color(self) -> str:
        """
        获取左边框颜色
        :return: 左边框颜色
        :rtype: str
        """
        return self._border_left_color

    def set_border_left_color(self, border_left_color: str) -> __init__:
        """
        设置左边框颜色
        :param border_left_color: 左边框颜色
        :type border_left_color: str
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """
        self._border_left_color = border_left_color
        return self

    @property
    def get_border_right_color(self) -> str:
        """
        获取右边框颜色
        :return: 右边框颜色
        :rtype: str
        """
        return self._border_right_color

    def set_border_right_color(self, border_right_color: str) -> __init__:
        """
        设置右边框颜色
        :param border_right_color: 右边框颜色
        :type border_right_color: str
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """
        self._border_right_color = border_right_color
        return self

    @property
    def get_border(self) -> openpyxl.styles.Border:
        """
        获取边框样式
        :return: 边框样式
        :rtype:
        """
        return openpyxl.styles.Border(
            top=openpyxl.styles.Side(style=self.get_border_top_style if self.get_border_top_style is not None else self.get_border_style, color=self.get_border_top_color if self.get_border_top_color is not None else self.get_border_color),
            bottom=openpyxl.styles.Side(style=self.get_border_bottom_style if self.get_border_bottom_style is not None else self.get_border_style, color=self.get_border_bottom_color if self.get_border_bottom_color is not None else self.get_border_color),
            left=openpyxl.styles.Side(style=self.get_border_left_style if self.get_border_left_style is not None else self.get_border_style, color=self.get_border_left_color if self.get_border_left_color is not None else self.get_border_color),
            right=openpyxl.styles.Side(style=self.get_border_right_style if self.get_border_right_style is not None else self.get_border_style, color=self.get_border_right_color if self.get_border_right_color is not None else self.get_border_color),
        )

    @property
    def get_fill_fg_color(self) -> str or None:
        """
        获取填充色
        :return: 填充色
        :rtype: str or None
        """
        return self._fill_fg_color

    def set_fill_fg_color(self, fill_fg_color) -> __init__:
        """
        设置填充色
        :param fill_fg_color: 填充色
        :type fill_fg_color: str
        :return: 本类对象
        :rtype: str
        """
        self._fill_fg_color = fill_fg_color
        return self

    @property
    def get_fill(self) -> openpyxl.styles.PatternFill | None:
        """
        获取填充样式
        :return: 填充样式
        :rtype: openpyxl.styles.PatternFill | None
        """
        if self.get_fill_fg_color:
            return openpyxl.styles.PatternFill(patternType='solid', fgColor=self.get_fill_fg_color)
        return None

    @property
    def get_alignment_horizontal(self) -> str:
        """
        获取字体水平样式
        :return: 字体水平样式
        :rtype: str
        """
        return self._alignment_horizontal

    def set_alignment_horizontal(self, alignment_horizontal: str) -> __init__:
        """
        设置字体水平样式
        :param alignment_horizontal: 字体水平样式
        :type alignment_horizontal: str
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterCell
        """
        self._alignment_horizontal = alignment_horizontal
        return self

    @property
    def get_alignment_vertical(self) -> str:
        """
        获取字体垂直样式
        :return: 字体垂直样式
        :rtype: str
        """
        return self._alignment_vertical

    def set_alignment_vertical(self, alignment_vertical: str) -> __init__:
        """
        设置字体垂直样式
        :param alignment_vertical: 字体垂直样式
        :type alignment_vertical:
        :return:
        :rtype:
        """
        self._alignment_vertical = alignment_vertical
        return self

    @property
    def get_alignment(self) -> openpyxl.styles.Alignment:
        """
        获取字体垂直、水平样式
        :return: 字体样式对象
        :rtype: openpyxl.styles.Alignment
        """
        return openpyxl.styles.Alignment(horizontal=self.get_alignment_horizontal, vertical=self.get_alignment_vertical)


class ExcelWriterRow:
    _cells: List[ExcelWriterCell] = []
    _row_index: int = None

    def __init__(self, row_index: int, excel_writer_cells: List[ExcelWriterCell]):
        self._row_index = row_index
        for column_index, excel_writer_cell in enumerate(excel_writer_cells):
            excel_writer_cell.set_coordinate(f'{openpyxl.utils.get_column_letter(column_index + 1)}{row_index}')
        self._cells = excel_writer_cells

    @property
    def get_row_index(self) -> int:
        """
        获取行索引
        :return:
        :rtype:
        """
        return self._row_index

    def set_row_index(self, row_index: int) -> __init__:
        """
        设置行索引
        :param row_index:
        :type row_index:
        :return:
        :rtype:
        """
        self._row_index = row_index
        return self

    @property
    def get_cells(self) -> List[ExcelWriterCell]:
        """
        获取单元格组
        :return: 若干单元格组成的字典
        :rtype: Dict[str:ExcelWriterCell]
        """
        return self._cells

    def set_cells(self, row_index: int, cells: List[ExcelWriterCell]) -> __init__:
        """
        设置单元格组
        :param row_index: 行索引
        :type row_index: int
        :param cells: 字典形式的单元格组合
        :type cells: List[ExcelWriterCell]
        :return: 本类对象
        :rtype: excelHelper.ExcelWriterRow
        """
        self._row_index = row_index
        for column_index, cells in enumerate(cells):
            cell.set_coordinate(f'{openpyxl.utils.get_column_letter(column_index + 1)}{row_index}')
        self._cells = cells
        return self


class ExcelWriter:
    _excel: openpyxl.Workbook = None
    _sheet: openpyxl.worksheet.worksheet.Worksheet = None
    _filename: str = ''

    def __init__(self, filename: str):
        """
        初始化
        :param filename: 文件名
        :type filename: str
        """
        self._excel = openpyxl.Workbook()
        self._filename = filename
        self._sheet = self._excel.active

    def __enter__(self) -> __init__:
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self._excel.close()

    def set_sheet_name(self, worksheet_name: str) -> __init__:
        """
        设置工作表名称
        :param worksheet_name: 工作表名称
        :type worksheet_name:  str
        :return: 本类对象
        :rtype: excelHelper.ExcelWriter
        """
        self._sheet.title = worksheet_name
        return self

    def add_sheet(self, worksheet_name: str, worksheet_index: int = None) -> __init__:
        """
        增加一个工作表
        :param worksheet_name: 工作簿名称
        :type worksheet_name: str
        :param worksheet_index: 工作簿所在位置
        :type worksheet_index: int
        :return: 本类对象
        :rtype: excelHelper.ExcelWriter
        """
        self._excel.create_sheet(title=worksheet_name, index=worksheet_index)
        self._sheet = self._excel[worksheet_name]
        return self

    def del_sheet(self, del_worksheet_name: str, select_worksheet_name: str = '') -> __init__:
        """
        删除一个工作表
        :param del_worksheet_name: 需要删除的工作表名称
        :type del_worksheet_name: str
        :param select_worksheet_name: 删除工作表之后，选中的工作表名称
        :type select_worksheet_name: str
        :return: 本类对象
        :rtype: excelHelper.ExcelWriter
        """
        del self._excel[del_worksheet_name]
        if select_worksheet_name:
            self._sheet = self._excel[select_worksheet_name]
        else:
            self._sheet = None
        return self

    def add_cell(self, excel_writer_cell: ExcelWriterCell) -> __init__:
        """
        添加单元格
        :param excel_writer_cell: 单元格对象
        :type excel_writer_cell: ExcelWriterCell
        :return: 本类对象
        :rtype: excelHelper.ExcelWriter
        """
        self._sheet[excel_writer_cell.get_coordinate] = excel_writer_cell.get_content
        self._sheet[excel_writer_cell.get_coordinate].font = excel_writer_cell.get_font
        self._sheet[excel_writer_cell.get_coordinate].border = excel_writer_cell.get_border
        if excel_writer_cell.get_fill is not None:
            self._sheet[excel_writer_cell.get_coordinate].fill = excel_writer_cell.get_fill
        self._sheet[excel_writer_cell.get_coordinate].alignment = excel_writer_cell.get_alignment
        return self

    def del_cell(self, location: str) -> __init__:
        """
        删除一个单元格
        :param location: 单元格坐标
        :type location: str
        :return: 本类对象
        :rtype: excelHelper.ExcelWriter
        """
        del self._sheet[location]
        return self

    def add_row(self, excel_writer_row: ExcelWriterRow) -> __init__:
        """
        添加行
        :param excel_writer_row: 行数据
        :type excel_writer_row: ExcelWriterRow
        :return: 本类对象
        :rtype: ExcelWriterCell
        """
        for excel_writer_cell in excel_writer_row.get_cells:
            self.add_cell(excel_writer_cell)
        return self

    def set_auto_filter_by_coordinate(self, original_cell_coordinate: str, finished_cell_coordinate: str) -> __init__:
        """
        通过坐标设置筛选范围
        :param original_cell_coordinate: 起始坐标
        :type original_cell_coordinate: str
        :param finished_cell_coordinate: 终止坐标
        :type finished_cell_coordinate: str
        :return: 本类对象
        :rtype: excelHelper.ExcelWriter
        """
        self._sheet.auto_filter.ref = f'{original_cell_coordinate}:{finished_cell_coordinate}'
        return self

    def set_auto_filter_by_column(self, column_index: int, values: []) -> __init__:
        """
        通过列索引设置筛选条件
        :param column_index: 列索引
        :type column_index: int
        :param values: 筛选条件
        :type values: str[]
        :return: 本类对象
        :rtype: excelHelper.ExcelWriter
        """
        self._sheet.auto_filter.add_filter_column(col_id=column_index, vars=values)
        return self

    def set_sort_condition(self, original_cell_coordinate: str, finished_cell_coordinate: str, descending: bool = False) -> __init__:
        """
        设置排序
        :param original_cell_coordinate: 起始坐标
        :type original_cell_coordinate: str
        :param finished_cell_coordinate: 终止坐标
        :type finished_cell_coordinate: str
        :param descending: 是否倒叙：False
        :type descending: bool
        :return: 本类对象
        :rtype: excelHelper.ExcelWriter
        """
        self._sheet.auto_filter.add_sort_condition(ref=f'{original_cell_coordinate}:{finished_cell_coordinate}', descending=descending)
        return self

    @property
    def get_max_column(self) -> int:
        """
        获取当前工作表列总数
        :return: 列总数
        :rtype: int
        """
        return self._sheet.max_column

    @property
    def get_max_row(self):
        """
        获取当前工作表行总数
        :return: 行总数
        :rtype: int
        """
        return self._sheet.max_row

    @property
    def get_row_dimensions(self) -> Dict[str, openpyxl.worksheet.dimensions.RowDimension]:
        """
        获取所有行高
        :return: 全表行高
        :rtype: dict
        """
        return self._sheet.row_dimensions

    def set_row_dimensions(self, row_index: int, row_height: float) -> __init__:
        """
        设置行高
        :param row_index: 行索引
        :type row_index: int
        :param row_height: 行高（0～409，默认12.75）
        :type row_height: float
        :return: 本类对象
        :rtype: excelHelper.ExcelWriter
        """
        self._sheet[row_index].height = row_height
        return self

    @property
    def get_column_dimensions(self) -> List[Tuple[str, openpyxl.worksheet.dimensions.ColumnDimension]]:
        """
        获取列宽
        :return: 全表列宽
        :rtype: list
        """
        return self._sheet.column_dimensions

    def set_column_dimensions(self, column_index: str, column_width: float) -> __init__:
        """
        设置列宽
        :param column_index: 列坐标
        :type column_index: str
        :param column_width: 列宽（0～255，默认：8.43个字符）
        :type column_width: float
        :return: 本类对象
        :rtype: excelHelper.ExcelWriter
        """
        self._sheet[column_index].width = column_width
        return self

    def set_merge_cells(self, original_cell_coordinate: str, finished_cell_coordinate: str) -> __init__:
        """
        设置合并单元格
        :param original_cell_coordinate: 起始坐标
        :type original_cell_coordinate: str
        :param finished_cell_coordinate: 终止坐标
        :type finished_cell_coordinate: str
        :return: 本类对象
        :rtype: excelHelper.ExcelWriter
        """
        self._sheet.merge_cells(f'{original_cell_coordinate}:{finished_cell_coordinate}')
        return self

    def set_unmerge_cells(self, original_cell_coordinate: str, finished_cell_coordinate: str) -> __init__:
        """
        拆分单元格
        :param original_cell_coordinate: 起始坐标
        :type original_cell_coordinate: str
        :param finished_cell_coordinate: 终止坐标
        :type finished_cell_coordinate: str
        :return: 本类对象
        :rtype: excelHelper.ExcelWriter
        """
        self._sheet.unmerge_cells(f'{original_cell_coordinate}:{finished_cell_coordinate}')
        return self

    def set_freeze_panes(self, cell_coordinate: str = None) -> __init__:
        """
        设置冻结窗口
        :param cell_coordinate: 单元格坐标（A1：冻结首行，B1：冻结首列，B2：冻结首行和首列，设置为None时取消冻结）
        :type cell_coordinate: str
        :return: 本类对象
        :rtype: excelHelper.ExcelWriter
        """
        self._sheet.freeze_panes = cell_coordinate
        return self

    def add_chart_bar(
            self,
            original_row: int,
            original_col: int,
            finished_row: int,
            finished_col: int,
            chart_target_coordinate: str,
            chart_title: str = None,
            x_axis_title: str = None,
            y_axis_title: str = None
    ):
        """
        添加条形图
        :param original_row:
        :type original_row:
        :param original_col:
        :type original_col:
        :param finished_row:
        :type finished_row:
        :param finished_col:
        :type finished_col:
        :param chart_target_coordinate:
        :type chart_target_coordinate:
        :param chart_title:
        :type chart_title:
        :param x_axis_title:
        :type x_axis_title:
        :param y_axis_title:
        :type y_axis_title:
        :return: 本类对象
        :rtype: excelHelper.ExcelWriter
        """
        values = openpyxl.chart.Reference(self._sheet, min_row=original_row, min_col=original_col, max_row=finished_row, max_col=finished_col)
        chart = openpyxl.chart.BarChart()
        if chart_title is not None:
            chart.title = chart_title
        if x_axis_title is not None:
            chart.x_axis.title = x_axis_title
        if y_axis_title is not None:
            chart.y_axis.title = y_axis_title
        chart.add_data(values, titles_from_data=True)

        # 设置分组
        x_label = openpyxl.chart.Reference(self._sheet, min_col=original_col, min_row=original_row, max_row=finished_row)
        chart.set_categories(x_label)

        self._sheet.add_chart(chart, chart_target_coordinate)
        return self

    def add_chart_line(
            self,
            original_row: int,
            original_col: int,
            finished_row: int,
            finished_col: int,
            chart_target_coordinate: str,
            chart_title: str = None,
            x_axis_title: str = None,
            y_axis_title: str = None
    ):
        """
        设置折线图
        :param original_row:
        :type original_row:
        :param original_col:
        :type original_col:
        :param finished_row:
        :type finished_row:
        :param finished_col:
        :type finished_col:
        :param chart_target_coordinate:
        :type chart_target_coordinate:
        :param chart_title:
        :type chart_title:
        :param x_axis_title:
        :type x_axis_title:
        :param y_axis_title:
        :type y_axis_title:
        :return: 本类对象
        :rtype: excelHelper.ExcelWriter
        """
        values = openpyxl.chart.Reference(self._sheet, min_row=original_row, min_col=original_col, max_row=finished_row, max_col=finished_col)
        chart = openpyxl.chart.LineChart()
        if chart_title is not None:
            chart.title = chart_title
        if x_axis_title is not None:
            chart.x_axis.title = x_axis_title
        if y_axis_title is not None:
            chart.y_axis.title = y_axis_title
        chart.add_data(values, titles_from_data=True)
        self._sheet.add_chart(chart, chart_target_coordinate)
        return self

    def add_chart_scatter(
            self,
            original_row: int,
            original_col: int,
            finished_row: int,
            finished_col: int,
            chart_target_coordinate: str,
            chart_title: str = None,
            x_axis_title: str = None,
            y_axis_title: str = None
    ):
        """
        添加散点图
        :param original_row:
        :type original_row:
        :param original_col:
        :type original_col:
        :param finished_row:
        :type finished_row:
        :param finished_col:
        :type finished_col:
        :param chart_target_coordinate:
        :type chart_target_coordinate:
        :param chart_title:
        :type chart_title:
        :param x_axis_title:
        :type x_axis_title:
        :param y_axis_title:
        :type y_axis_title:
        :return: 本类对象
        :rtype: excelHelper.ExcelWriter
        """
        values = openpyxl.chart.Reference(self._sheet, min_row=original_row, min_col=original_col, max_row=finished_row, max_col=finished_col)
        chart = openpyxl.chart.ScatterChart()
        if chart_title is not None:
            chart.title = chart_title
        if x_axis_title is not None:
            chart.x_axis.title = x_axis_title
        if y_axis_title is not None:
            chart.y_axis.title = y_axis_title
        chart.add_data(values, titles_from_data=True)
        self._sheet.add_chart(chart, chart_target_coordinate)
        return self

    def add_chart_pie(
            self,
            original_row: int,
            original_col: int,
            finished_row: int,
            finished_col: int,
            chart_target_coordinate: str,
            chart_title: str = None,
            x_axis_title: str = None,
            y_axis_title: str = None
    ):
        """
        添加饼图
        :param original_row:
        :type original_row:
        :param original_col:
        :type original_col:
        :param finished_row:
        :type finished_row:
        :param finished_col:
        :type finished_col:
        :param chart_target_coordinate:
        :type chart_target_coordinate:
        :param chart_title:
        :type chart_title:
        :param x_axis_title:
        :type x_axis_title:
        :param y_axis_title:
        :type y_axis_title:
        :return: 本类对象
        :rtype: excelHelper.ExcelWriter
        """
        values = openpyxl.chart.Reference(self._sheet, min_row=original_row, min_col=original_col, max_row=finished_row, max_col=finished_col)
        chart = openpyxl.chart.PieChart()
        if chart_title is not None:
            chart.title = chart_title
        if x_axis_title is not None:
            chart.x_axis.title = x_axis_title
        if y_axis_title is not None:
            chart.y_axis.title = y_axis_title
        chart.add_data(values, titles_from_data=True)
        self._sheet.add_chart(chart, chart_target_coordinate)
        return self

    def save(self) -> __init__:
        """
        保存文件
        :return:
        :rtype:
        """
        self._excel.save(self._filename)
        return self


if __name__ == '__main__':
    """
    参数说明
    -T --operation_type：【三个可选参数reader、writer和update】分别控制演示读取、写入和修改原表三个功能
    """
    parser = argparse.ArgumentParser()
    parser.description = 'excel读取文件工区（只支持2007版）'
    parser.add_argument('-T', '--operation_type', help='执行类型：reader、writer、update', type=str, default='')
    args = parser.parse_args()
    operation_type = args.operation_type

    input_filename = 'factories.xlsx'
    relative_path = True

    if operation_type == 'reader':
        # ExcelReader演示
        with ExcelReader(filename=os.path.join(sys.path[0], input_filename) if relative_path else input_filename) as xlrd:
            # 通过工作表名称进行读取
            excel_content = xlrd.read_entire_sheet_by_name('Sheet1').to_dict
            print('example1：', excel_content)

            # 读取第一个Sheet
            excel_content = xlrd.read_entire_sheet_by_first().to_dict
            print('example：', excel_content)

            # 通过激活Sheet进行读取
            excel_content = xlrd.read_entire_sheet_by_active().to_dict
            print('example3：', excel_content)

            # 读取表头
            excel_title = xlrd.set_sheet_by_active().set_read_title_row_number(1).read_title().get_title
            print('example：', excel_title)

            # 读取多行数据
            excel_content = xlrd.set_sheet_by_active().set_original_row_number(2).set_finished_row_number(5).set_title(excel_title).read_rows().to_dict
            print('example5：', excel_content)

            # 获取列表数据
            excel_content = xlrd.set_sheet_by_active().set_original_row_number(2).set_finished_row_number(5).set_title(excel_title).read_rows().to_list
            print('example6：', excel_content)

            print(xlrd.set_sheet_by_active().get_row_dimensions)

    elif operation_type == 'writer':
        # ExcelWriter演示
        # 设置单独单元格写入
        with ExcelWriter(filename=os.path.join(sys.path[0], 'test-1.xlsx') if relative_path else 'test-1.xlsx') as xlwt:
            xlwt.set_sheet_name('测试表'). \
                add_cell(
                ExcelWriterCell(
                    content='测试1',  # 单元格内容
                    coordinate='A1',  # 单元格坐标
                    font_name='宋体',  # 字体
                    font_size=18,  # 字号
                    font_color='FF0000',  # 字体颜色
                    alignment_vertical=ExcelWriterCell.ALIGNMENT_VERTICAL_TOP,  # 垂直样式
                    alignment_horizontal=ExcelWriterCell.ALIGNMENT_HORIZONTAL_RIGHT,  # 水平样式
                )
            ). \
                add_cell(
                ExcelWriterCell(
                    content='测试2',  # 单元格内容
                    coordinate='B2',  # 单元格坐标
                    border_style=ExcelWriterCell.BORDER_STYLE_DOTTED,  # 整体边框样式
                    border_color='00FFFF',  # 整体边框颜色
                    border_left_style=ExcelWriterCell.BORDER_STYLE_MEDIUM_DASH_DOT,  # 左边框样式
                    fill_fg_color='FFFF00',  # 填充背景色
                )
            ). \
                add_cell(
                ExcelWriterCell(
                    content='合并测试',
                    coordinate='C1',
                    alignment_horizontal=ExcelWriterCell.ALIGNMENT_HORIZONTAL_CENTER,
                    alignment_vertical=ExcelWriterCell.ALIGNMENT_VERTICAL_CENTER
                )
            ). \
                set_auto_filter_by_coordinate(original_cell_coordinate='A1', finished_cell_coordinate=f'B{xlwt.get_max_row}'). \
                set_merge_cells('C1', 'D5'). \
                set_freeze_panes('B2'). \
                save()

        # 设置一行单元格写入
        with ExcelWriter(filename=os.path.join(sys.path[0], 'test-2.xlsx') if relative_path else 'test-2.xlsx') as xlwt:
            # 定义表头
            xlwt.add_row(
                ExcelWriterRow(row_index=1, excel_writer_cells=[
                    ExcelWriterCell(content='姓名'),
                    ExcelWriterCell(content='工龄'),
                    ExcelWriterCell(content='销售额'),
                ])
            )

            # 定义销售表单
            data = [
                {'姓名': '张三', '工龄': 5, '销售额': 5000, },
                {'姓名': '李四', '工龄': 7, '销售额': 7000, },
                {'姓名': '王五', '工龄': 6, '销售额': 6000, },
                {'姓名': '赵六', '工龄': 10, '销售额': 8888, },
            ]
            for idx, datum in enumerate(data):
                a, b, c = datum.values()
                xlwt.add_row(
                    ExcelWriterRow(row_index=idx + 2, excel_writer_cells=[
                        ExcelWriterCell(content=a).set_font_color(font_color='FF0000', when=True),
                        ExcelWriterCell(content=b).set_font_color(font_color='00FF00'),
                        ExcelWriterCell(content=c, font_color='0000FF'),
                    ])
                )
            xlwt.save()

        # 制作Excel图表
        with ExcelWriter(filename=os.path.join(sys.path[0], 'test-3.xlsx') if relative_path else 'test-3.xlsx') as xlwt:
            xlwt. \
                add_cell(ExcelWriterCell(content='姓名', coordinate='A1', )). \
                add_cell(ExcelWriterCell(content='工资', coordinate='B1', )). \
                add_cell(ExcelWriterCell(content='张三', coordinate='A2', )). \
                add_cell(ExcelWriterCell(content=100, coordinate='B2', )). \
                add_cell(ExcelWriterCell(content='李四', coordinate='A3', )). \
                add_cell(ExcelWriterCell(content=200, coordinate='B3', )). \
                add_cell(ExcelWriterCell(content='王五', coordinate='A4', )). \
                add_cell(ExcelWriterCell(content=300, coordinate='B4', )). \
                add_cell(ExcelWriterCell(content='赵六', coordinate='A5', )). \
                add_cell(ExcelWriterCell(content=400, coordinate='B5', )). \
                add_chart_bar(original_row=2, original_col=1, finished_row=5, finished_col=2, chart_target_coordinate='F1'). \
                add_chart_line(original_row=2, original_col=1, finished_row=5, finished_col=2, chart_target_coordinate='G2'). \
                add_chart_scatter(original_row=2, original_col=1, finished_row=5, finished_col=2, chart_target_coordinate='H3'). \
                add_chart_pie(original_row=2, original_col=1, finished_row=5, finished_col=2, chart_target_coordinate='I4'). \
                save()


    elif operation_type == 'update':
        # 修改表格
        PRO_PRICES = {'大蒜': 1, '芹菜': 2, '芒果': 3, }
        with ExcelReader(filename=os.path.join(sys.path[0], '工作簿1.xlsx') if relative_path else '工作簿1.xlsx') as xlrd:
            for row in xlrd.set_sheet_by_index().read_rows().get_rows:
                if row[0].value in PRO_PRICES:
                    cell = ExcelWriterCell(
                        content=PRO_PRICES[row[0].value],
                        font_name='宋体',
                        font_size=18,
                        font_bold=True,
                        font_underline='single',
                        font_color='00FF00',
                        border_style=ExcelWriterCell.BORDER_STYLE_MEDIUM,
                        border_color='FFFF00',
                    )
                    row[1].value = cell.get_content
                    row[1].font = cell.get_font
                    row[1].border = cell.get_border

            xlrd.save()
